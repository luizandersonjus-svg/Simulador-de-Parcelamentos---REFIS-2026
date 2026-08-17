from __future__ import annotations
import io
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from parsers import PLAN_LABELS


def dataframe_to_xlsx(df: pd.DataFrame, sheet_name: str = "Resultado",
                      flags: list[dict] | None = None) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in range(1, ws.max_column + 1):
            values = [str(ws.cell(row, column).value or "") for row in range(1, min(ws.max_row, 100) + 1)]
            ws.column_dimensions[get_column_letter(column)].width = min(max(max(map(len, values)) + 2, 12), 42)
        money_columns = ["Normal"] + [label for label, _ in PLAN_LABELS if label in df.columns]
        for col in money_columns:
            idx = list(df.columns).index(col) + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = 'R$ #,##0.00'
        if ws.max_row > 1 and str(ws.cell(row=ws.max_row, column=1).value or "").strip().upper() == "TOTAL":
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E2F3")
        if flags:
            for flag in flags:
                col, row = flag.get("column"), flag.get("row")
                if (col not in df.columns or not isinstance(row, int)
                        or row < 0 or row >= len(df)):
                    continue
                cell = ws.cell(row=row + 2, column=list(df.columns).index(col) + 1)
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(color="9C0006", bold=True)
                cell.comment = Comment(flag.get("message", ""), "Sistema", height=90, width=320)
    return buffer.getvalue()
